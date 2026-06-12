#!/usr/bin/env python3
"""
Muse Video Skill — export_xlsx.py
Input:  Project State JSON (stdin or --input file)
Output: Excel workbook with storyboard tech breakdown + project summary

Role: One job — render Project State → formatted Excel (.xlsx).
      Reads the schema from assets/templates/export/script-tech.xlsx (JSON format).
      Requires openpyxl (pip install openpyxl).

Schema-driven: reads column definitions from the template JSON, no hardcoded layout.
"""

import json
import sys
import argparse
from datetime import datetime, timezone
from pathlib import Path

VERSION = "0.3.0"


def safe_str(val, default: str = "—") -> str:
    if val is None:
        return default
    s = str(val).strip()
    return s if s else default


def safe_list(val) -> list:
    if isinstance(val, list):
        return val
    return []


def resolve_path(data: dict, dotted_path: str, default=None):
    keys = dotted_path.split(".")
    current = data
    for k in keys:
        if isinstance(current, dict):
            current = current.get(k)
        else:
            return default
        if current is None:
            return default
    return current


def build_xlsx(project_state: dict, schema: dict, output_path: str) -> None:
    """Generate Excel workbook from Project State + schema definition."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(
            "ERROR: openpyxl is required. Install with: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = Workbook()

    # Global styles
    gs = schema.get("global_style", {})
    font_name = gs.get("font_name", "Microsoft YaHei")
    font_size = gs.get("font_size", 10)
    header_fill_hex = gs.get("header_fill", "cc785c")
    header_font_color = gs.get("header_font_color", "ffffff")
    border_style = gs.get("border_style", "thin")

    thin_border = Border(
        left=Side(style=border_style, color="cccccc"),
        right=Side(style=border_style, color="cccccc"),
        top=Side(style=border_style, color="cccccc"),
        bottom=Side(style=border_style, color="cccccc"),
    )
    header_fill = PatternFill(start_color=header_fill_hex, end_color=header_fill_hex, fill_type="solid")
    header_font = Font(name=font_name, size=gs.get("header_font_size", 11),
                       bold=True, color=header_font_color)
    body_font = Font(name=font_name, size=font_size)

    for sheet_def in schema.get("sheets", []):
        sheet_name = sheet_def.get("name", "Sheet")
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31 char limit

        # ── Column-based sheet ──
        if "columns" in sheet_def:
            columns = sheet_def["columns"]
            data_key = sheet_def.get("data_key", "")
            data_rows = []
            if data_key:
                data_rows = safe_list(resolve_path(project_state, data_key))
            else:
                # Derive data from storyboard
                storyboard = project_state.get("storyboard", [])
                script_scenes = resolve_path(project_state, "script.scenes", [])
                for panel in storyboard:
                    row = dict(panel)
                    # Merge scene info
                    scene = next(
                        (s for s in safe_list(script_scenes)
                         if s.get("scene_id") == panel.get("scene_id")),
                        {},
                    )
                    row["scene_title"] = scene.get("scene_title", "")
                    row["duration"] = scene.get("duration", "")
                    cam = panel.get("camera", {}) if isinstance(panel.get("camera"), dict) else {}
                    row["shot_type"] = cam.get("shot_type", "") if isinstance(cam, dict) else ""
                    row["movement"] = cam.get("movement", "") if isinstance(cam, dict) else ""
                    row["lens"] = cam.get("lens", "") if isinstance(cam, dict) else ""
                    row["lighting"] = cam.get("lighting", "") if isinstance(cam, dict) else ""
                    row["sound_notes"] = scene.get("sound_notes", "")
                    meta = panel.get("_meta", {})
                    row["revision"] = meta.get("revision", 1) if isinstance(meta, dict) else 1
                    char_ids = panel.get("character_ids", [])
                    row["character_ids"] = ", ".join(str(c) for c in char_ids) if char_ids else ""
                    data_rows.append(row)

            # Write header row
            for ci, col in enumerate(columns, 1):
                cell = ws.cell(row=1, column=ci, value=col.get("header", ""))
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[get_column_letter(ci)].width = col.get("width", 15)

            # Write data rows
            for ri, row_data in enumerate(data_rows, 2):
                for ci, col in enumerate(columns, 1):
                    key = col.get("key", "")
                    val = row_data.get(key, "")
                    if isinstance(val, bool):
                        val = "✓ 已审核" if val else "⏳ 待审核"
                    elif isinstance(val, list):
                        val = ", ".join(str(x) for x in val)
                    elif val is None:
                        val = ""
                    cell = ws.cell(row=ri, column=ci, value=str(val))
                    cell.font = body_font
                    cell.border = thin_border

            # Freeze panes
            freeze = sheet_def.get("freeze_panes", "A2")
            ws.freeze_panes = freeze

            # Conditional formatting
            cf = sheet_def.get("conditional_formatting", {})
            for rule_name, rule in cf.items():
                col_key = rule.get("column", "")
                col_idx = None
                for ci, col in enumerate(columns, 1):
                    if col.get("key") == col_key:
                        col_idx = ci
                        break
                if col_idx:
                    true_fill = PatternFill(start_color=rule.get("fill", "d4edda"),
                                            end_color=rule.get("fill", "d4edda"),
                                            fill_type="solid")
                    for ri in range(2, len(data_rows) + 2):
                        cell = ws.cell(row=ri, column=col_idx)
                        if cell.value and "已审核" in str(cell.value):
                            cell.fill = true_fill

        # ── Row-based sheet (key-value pairs) ──
        elif "rows" in sheet_def:
            rows = sheet_def["rows"]
            label_font = Font(name=font_name, size=font_size, bold=True)
            label_fill = PatternFill(start_color="f5f3ef", end_color="f5f3ef", fill_type="solid")
            col_a_width = sheet_def.get("label_style", {}).get("column_width_a", 16)
            col_b_width = sheet_def.get("label_style", {}).get("column_width_b", 50)
            ws.column_dimensions["A"].width = col_a_width
            ws.column_dimensions["B"].width = col_b_width

            for ri, row_def in enumerate(rows, 1):
                label = row_def.get("label", "")
                key = row_def.get("key", "")
                val = resolve_path(project_state, key, "")

                # Special handling for lists
                if isinstance(val, list) and key.endswith(("_refs", "_notes", "_prompts")):
                    val = "\n".join(f"• {v}" for v in val)

                cell_label = ws.cell(row=ri, column=1, value=label)
                cell_label.font = label_font
                cell_label.fill = label_fill
                cell_label.border = thin_border

                cell_val = ws.cell(row=ri, column=2, value=str(val) if val else "")
                cell_val.font = body_font
                cell_val.border = thin_border
                cell_val.alignment = Alignment(wrap_text=True)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Set first sheet as active
    if wb.sheetnames:
        wb.active = wb.sheetnames.index(wb.sheetnames[0])

    wb.save(output_path)
    print(f"✅ Excel workbook → {output_path}", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(
        description="Muse Video Skill — Export Project State → Excel (.xlsx)"
    )
    parser.add_argument("--input", "-i", type=str, default=None,
                        help="Path to Project State JSON (default: stdin)")
    parser.add_argument("--output", "-o", type=str, required=True,
                        help="Output .xlsx file path")
    parser.add_argument("--schema", type=str, default=None,
                        help="Path to Excel schema JSON (default: assets/templates/export/script-tech.xlsx)")
    args = parser.parse_args()

    # Load schema
    if args.schema:
        schema_path = args.schema
    else:
        script_dir = Path(__file__).resolve().parent.parent
        schema_path = str(script_dir / "assets" / "templates" / "export" / "script-tech.xlsx")
    with open(schema_path, "r", encoding="utf-8") as f:
        schema = json.load(f)

    # Load Project State
    if args.input:
        with open(args.input, "r", encoding="utf-8") as f:
            project_state = json.load(f)
    else:
        project_state = json.load(sys.stdin)

    # Build XLSX
    build_xlsx(project_state, schema, args.output)


if __name__ == "__main__":
    main()
